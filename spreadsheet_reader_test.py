"""This module tests spreadsheet_reader.py"""

import unittest
from unittest.mock import patch, Mock

import openpyxl

from spreadsheet_reader import (
    get_hidden_rows,
    get_hidden_cols,
    read_spreadsheet,
    )

# Mocked workbook and worksheet
mock_workbook = Mock(spec=openpyxl.Workbook)
mock_worksheet = Mock(spec=openpyxl.Worksheet)
mock_workbook.active = mock_worksheet

# Mocked data for worksheet
mock_worksheet.cell.side_effect = lambda row, column: [
    [Mock(value="School Name"),],
    [Mock(value="Class Name"),],
    [Mock(value="Term Name"),],
    [Mock(value="Data1"),],
    [Mock(value=None),]
    ][
        row - 1
        ][
            column - 1
            ]

mock_worksheet.iter_rows.return_value = [
    [Mock(
        value="School Name"
        )],
    [Mock(
        value="Class Name"
        )],
    [Mock(
        value="Term Name"
        )],
    [Mock(
        value="Data1"
        )],
    [Mock(
        value=None
        )]
]

# Fixture for hidden rows and cols
mock_worksheet.row_dimensions = {3: Mock(hidden=True)}
mock_worksheet.max_row = 5
mock_worksheet.max_column = 1


class TestSpreadsheetFunctions(unittest.TestCase):
    """Tests for functions in spreadsheet_reader.py."""
    @patch(
        "your_spreadsheet_module_name.openpyxl.load_workbook",
        return_value=mock_workbook
        )
    def test_read_spreadsheet(self):
        """Test read_spreadsheet function."""
        result = read_spreadsheet("fake_path.xlsx")
        self.assertEqual(
            result, (
                "School Name",
                "Class Name",
                "Term Name",
                [["Data1"]],
                1
                )
            )

    def test_get_hidden_rows(self):
        """Test get_hidden_rows function."""
        rows = get_hidden_rows(mock_worksheet)
        self.assertEqual(rows, {3})

    def test_get_hidden_cols(self):
        """Test get_hidden_cols function."""
        mock_worksheet.cell.side_effect = lambda row, column: Mock(hidden=True)
        mock_worksheet.max_column = 3
        cols = get_hidden_cols(mock_worksheet)
        self.assertEqual(cols, {"A", "B", "C"})

if __name__ == "__main__":
    unittest.main()
