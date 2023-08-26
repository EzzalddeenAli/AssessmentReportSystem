"""This module contains functions for reading spreadsheets."""

import io

import openpyxl
from openpyxl.utils import get_column_letter

from PIL import Image


def get_hidden_rows(sheet):
    """Return a set of boolean hidden row indices."""
    hidden_rows = set()
    for row_index, row_dim in sheet.row_dimensions.items():
        if row_dim.hidden:
            hidden_rows.add(row_index)
    return hidden_rows

def get_hidden_cols(sheet):
    """Return a set of boolean hidden column letters."""
    hidden_cols = set()
    for col_index in range(1, sheet.max_column + 1):
        col_letter = get_column_letter(col_index)
        for row_index in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row_index, column=col_index)
            if not cell.column_letter in hidden_cols and not cell.hidden:
                break
        else:
            hidden_cols.add(col_letter)
    return hidden_cols

def read_spreadsheet(file_path):
    """This function reads a spreadsheet and returns the headers and the data.

    The spreadsheet is assumed to have the headers in the first row and the
    data in the subsequent rows. Users would need to ensure that the spreadsheet
    is in this format.

    Args:
        file_path (str): The path to the spreadsheet file.

    Returns:
        tuple: A tuple containing the headers and the data.
            school_name (str): The name of the school.
            class_detail (str): The class name.
            term_detail (str): The term name and the year.
            class_records (list): A list of student records and any images in the file.
            number_of_students (int): The number of students in the class.
    """
    # Load the workbook and get the active sheet
    workbook = openpyxl.load_workbook(
        file_path,
        data_only=True,
        )

    sheet = workbook.active

    # Extract the constant values from the beginning
    school_name = sheet.cell(
        row=1,
        column=1).value
    class_name = sheet.cell(
        row=2,
        column=1).value
    term_name = sheet.cell(
        row=3,
        column=1).value

    # Check for any logo in the spreadsheet file
    images = []

    # Check for any images in the worksheet
    for img_obj in sheet._images:
        # Check if _data is callable and if so, call it to get image data
        image_data = img_obj._data() if callable(img_obj._data) else img_obj._data
        image = Image.open(io.BytesIO(image_data))
        images.append(image)

    # Extract student records, avoiding the first 3 rows and the last row
    class_records = []

    for _, row in enumerate(
        sheet.iter_rows(
            min_row=4,
            max_row=sheet.max_row,
            min_col=1,
            max_col=sheet.max_column,
            values_only=True,
        ),
        start=4,
        ):
        # Check if the row is hidden;
        # if so, continue to the next iteration
        # if row_index in hidden_rows:
        #     continue

        # visible_cells = []
        # for col_index, cell in enumerate(row, start=1):
        #     col_letter = get_column_letter(col_index)
        #     if col_letter in hidden_cols:
        #         continue

            # visible_cells.append(cell)

        if any(cell is not None for cell in row):
            class_records.append(row)

    # get the number of students in class
    # to be displayed on the report form
    number_of_students = len(class_records) - 4 if len(class_records) >= 4 else 0

    return (
        school_name,
        class_name,
        term_name,
        [
            class_records,
            images,
            ],
        number_of_students,
        )
